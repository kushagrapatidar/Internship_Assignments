import openpyxl
from openpyxl.styles import Font, Alignment

# Save the Data to 
def save_data(data, fname):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a new sheet in the workbook
    sheet = workbook.active

    # Add the dictionary data to the Excel sheet
    header = list(data.keys())
    sheet.append(header)  # Add the header row
    
    for row in zip(*data.values()):
        sheet.append(row)

    # Save the workbook to an XLSX file
    workbook.save(fname)

# Reformat the xlsx file to Output Data Structure
def reformat(fname):
        # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(fname)

    # Select a specific sheet in the workbook
    sheet = workbook.active

    font = Font(name='Calibri', bold=True, color='000000')
    alignment = Alignment(horizontal='left', vertical='center')

    # Define the range of cells to reformat
    start_row, end_row = 1, 1
    start_column, end_column = 1, 15

    # Apply the font and alignment to the specified range of cells
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    font = Font(name='Calibri', bold=False, color='000000')
    alignment = Alignment(horizontal='right', vertical='center')
    start_row, end_row = 2, 115
    start_column, end_column = 1, 1

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    font = Font(name='Calibri', bold=False, color='000000')
    alignment = Alignment(horizontal='right', vertical='center')
    start_row, end_row = 2, 115
    start_column, end_column = 3, 15

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # Autofit the coulmn width
    col_num = 0
    for column in sheet.columns:
        adjusted_width = 18.11

        if col_num == 0:
            adjusted_width = 7.89
        if col_num == 1:
            adjusted_width = 119.56

        sheet.column_dimensions[chr(col_num+65)].width = adjusted_width
        col_num += 1

    # Save the modified workbook
    workbook.save(fname)
